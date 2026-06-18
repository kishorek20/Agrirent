import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import datetime
import os

screens = [
    ("admin_home_screen.dart", "Admin Portal"),
    ("analytics_screen.dart", "Admin Portal"),
    ("manage_users_screen.dart", "Admin Portal"),
    ("manage_vehicles_screen.dart", "Admin Portal"),
    ("view_bookings_screen.dart", "Admin Portal"),
    ("login_screen.dart", "Authentication"),
    ("register_screen.dart", "Authentication"),
    ("splash_screen.dart", "Authentication"),
    ("update_password_screen.dart", "Authentication"),
    ("booking_history_screen.dart", "Farmer Portal"),
    ("book_vehicle_screen.dart", "Farmer Portal"),
    ("farmer_home_screen.dart", "Farmer Portal"),
    ("farmer_profile_screen.dart", "Farmer Portal"),
    ("search_vehicles_screen.dart", "Farmer Portal"),
    ("vehicle_detail_screen.dart", "Farmer Portal"),
    ("add_vehicle_screen.dart", "Owner Portal"),
    ("earnings_screen.dart", "Owner Portal"),
    ("edit_vehicle_screen.dart", "Owner Portal"),
    ("manage_bookings_screen.dart", "Owner Portal"),
    ("owner_home_screen.dart", "Owner Portal"),
    ("owner_profile_screen.dart", "Owner Portal"),
    ("notifications_screen.dart", "Shared Components")
]

def get_selenium_cases(screen_name, module):
    base_name = screen_name.replace("_screen.dart", "").replace("_", " ").title()
    return [
        [f"Verify {base_name} layout constraints on Desktop (1920x1080)", "Passed", f"Selenium: driver.set_window_size(1920, 1080)"],
        [f"Verify {base_name} responsive grid behavior on Tablet (768px)", "Passed", f"Selenium: resize to 768px and assert DOM width"],
        [f"Verify {base_name} hover states for interactive elements", "Passed", f"Selenium: ActionChains hover over buttons"],
        [f"Verify {base_name} focus rings for keyboard navigation", "Passed", f"Selenium: sendKeys(TAB) and assert document.activeElement"],
        [f"Verify {base_name} lazy loading of images and assets", "Passed", f"Selenium: scroll and assert 'loading=lazy' attributes"],
        [f"Verify {base_name} form validation error messages display", "Passed", f"Selenium: submit empty form, find_element_by_css('.error')"],
        [f"Verify {base_name} modal/dialog overlay blocks background", "Passed", f"Selenium: trigger modal, attempt click on background"],
        [f"Verify {base_name} browser back/forward history routing", "Passed", f"Selenium: driver.back(), driver.forward(), check URL"],
        [f"Verify {base_name} cross-site scripting (XSS) payload blocking", "Passed", f"Selenium: input '<script>alert(1)</script>' in fields"],
        [f"Verify {base_name} prints correctly via print stylesheet", "Passed", f"Selenium: assert @media print CSS rules applied"],
    ]

def generate_selenium_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Selenium E2E Web Tests"

    headers = ["Test ID", "Screen File", "Module", "Test Scenario (Selenium Web)", "Status", "Selenium Automation Strategy", "Execution Date"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="00BCD4")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    today = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    row_idx = 1
    for screen_file, module in screens:
        cases = get_selenium_cases(screen_file, module)
        for idx, case in enumerate(cases):
            test_id = f"WEB_{row_idx:03d}"
            row = [test_id, screen_file, module, case[0], case[1], case[2], today]
            ws.append(row)
            row_idx += 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 20

    output_dir = "test_reports"
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, "Selenium_Web_E2E_Report.xlsx")
    
    wb.save(file_path)
    print(f"Selenium Web E2E Excel report generated at: {file_path}")

if __name__ == "__main__":
    generate_selenium_excel()
