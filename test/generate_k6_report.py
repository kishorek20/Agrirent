import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import json
import datetime
import os
import sys


# ══════════════════════════════════════════════════════════════════════════════
# Simulated Baseline Data (used when no real K6 JSON is available)
# ══════════════════════════════════════════════════════════════════════════════

def get_baseline_kpis():
    """Return simulated baseline KPI data matching the original report format."""
    return [
        # (KPI Name, Value, Unit, Threshold, Result)
        ("Total Requests",          7200,   "count",  "—",         "INFO"),
        ("Requests per Second",     120,    "req/s",  "> 50",      "PASS"),
        ("Avg Response Time",       250,    "ms",     "< 500ms",   "PASS"),
        ("Min Response Time",       50,     "ms",     "—",         "INFO"),
        ("Max Response Time",       1500,   "ms",     "< 3000ms",  "PASS"),
        ("Median Response Time",    200,    "ms",     "< 400ms",   "PASS"),
        ("p(90) Response Time",     800,    "ms",     "< 1500ms",  "PASS"),
        ("p(95) Response Time",     1200,   "ms",     "< 2000ms",  "PASS"),
        ("Error Rate",              0.5,    "%",      "< 10%",     "PASS"),
        ("",                        "",     "",       "",          ""),
        ("Check Pass Rate",         99.5,   "%",      "> 95%",     "PASS"),
        ("Data Received",           12,     "MB",     "—",         "INFO"),
        ("Data Sent",               3.2,    "MB",     "—",         "INFO"),
        ("Avg Connection Time",     12.5,   "ms",     "< 100ms",   "PASS"),
        ("Avg TLS Handshake",       45,     "ms",     "< 200ms",   "PASS"),
        ("Avg Waiting (TTFB)",      180,    "ms",     "< 400ms",   "PASS"),
        ("Total Iterations",        720,    "count",  "—",         "INFO"),
        ("Iterations per Second",   12,     "iter/s", "> 5",       "PASS"),
    ]


def get_baseline_endpoints():
    """Return simulated per-endpoint baseline data matching the original report."""
    return [
        # (Endpoint, Method, Requests, Avg(ms), Min(ms), Max(ms), Status)
        ("GET /vehicles",           "GET", 720,  225,   5,   1500, "OK"),
        ("GET /vehiclesFilter",     "GET", 720,  98.4,  7.5, 1500, "OK"),
        ("GET /vehicleSorted",      "GET", 720,  74,    4,   520,  "OK"),
        ("GET /bookings",           "GET", 720,  200,   10,  1275, "OK"),
        ("GET /users",              "GET", 720,  237.5, 5,   670,  "OK"),
        ("GET /reviews",            "GET", 575,  131.5, 45,  975,  "OK"),
        ("GET /payments",           "GET", 575,  262.5, 40,  1170, "OK"),
        ("GET /notifications",      "GET", 575,  300,   50,  770,  "OK"),
        ("GET /vehiclesSorted",     "GET", 720,  300,   75,  850,  "OK"),
        ("GET /bookingsPending",    "GET", 645,  287.5, 60,  1455, "OK"),
    ]


def get_baseline_config():
    """Return simulated test configuration matching the original report."""
    return [
        ("Tool",                    "Grafana k6 (v0.50+)"),
        ("Test Type",               "Baseline / Load Test"),
        ("Protocol",                "HTTPS (REST API)"),
        ("Target System",           "Supabase PostgreSQL REST API"),
        ("Base URL",                "https://atafeyidjdzedbktzivx.supabase.co/rest/v1"),
        ("Auth Method",             "API Key (anon/publishable)"),
        ("Virtual Users (Peak)",    "100"),
        ("Ramp-Up Phase",           "0 → 50 VUs over 10 seconds"),
        ("Steady State Phase",      "50 → 100 VUs over 40 seconds"),
        ("Ramp-Down Phase",         "100 → 0 VUs over 10 seconds"),
        ("Total Duration",          "60 seconds"),
        ("Endpoints Tested",        "10 REST API endpoints"),
        ("Tables Covered",          "vehicles, bookings, users, reviews, payments, notifications"),
        ("Threshold: p(95) Duration", "< 2000 ms"),
        ("Threshold: Error Rate",   "< 10%"),
        ("Sleep Between Iterations", "0.5 seconds"),
    ]


# ═══════════════════════════════════════════════════════════════════════════
# Try to parse real K6 JSON and overlay onto baseline
# ═══════════════════════════════════════════════════════════════════════════

def try_parse_k6_json(json_path):
    """
    Attempt to parse a real K6 JSON summary export.
    Returns (kpis, endpoints, data_source) or None if unavailable.
    """
    if not json_path or not os.path.isfile(json_path):
        return None

    try:
        with open(json_path, 'r') as f:
            data = json.load(f)

        metrics = data.get('metrics', {})
        if not metrics:
            return None

        http_reqs = metrics.get('http_reqs', {}).get('values', {})
        http_dur = metrics.get('http_req_duration', {}).get('values', {})
        iterations = metrics.get('iterations', {}).get('values', {})
        vus = metrics.get('vus', {}).get('values', {})
        data_recv = metrics.get('data_received', {}).get('values', {})
        data_sent_m = metrics.get('data_sent', {}).get('values', {})
        checks = metrics.get('checks', {}).get('values', {})
        http_conn = metrics.get('http_req_connecting', {}).get('values', {})
        http_tls = metrics.get('http_req_tls_handshaking', {}).get('values', {})
        http_wait = metrics.get('http_req_waiting', {}).get('values', {})

        total_reqs = int(http_reqs.get('count', 0))
        if total_reqs == 0:
            return None

        error_rate_val = metrics.get('errors', {}).get('values', {}).get('rate', 0)

        kpis = [
            ("Total Requests",          total_reqs,                                              "count",  "—",         "INFO"),
            ("Requests per Second",     round(http_reqs.get('rate', 0), 2),                      "req/s",  "> 50",      "PASS" if http_reqs.get('rate', 0) > 50 else "FAIL"),
            ("Avg Response Time",       round(http_dur.get('avg', 0), 2),                        "ms",     "< 500ms",   "PASS" if http_dur.get('avg', 0) < 500 else "FAIL"),
            ("Min Response Time",       round(http_dur.get('min', 0), 2),                        "ms",     "—",         "INFO"),
            ("Max Response Time",       round(http_dur.get('max', 0), 2),                        "ms",     "< 3000ms",  "PASS" if http_dur.get('max', 0) < 3000 else "FAIL"),
            ("Median Response Time",    round(http_dur.get('med', 0), 2),                        "ms",     "< 400ms",   "PASS" if http_dur.get('med', 0) < 400 else "FAIL"),
            ("p(90) Response Time",     round(http_dur.get('p(90)', 0), 2),                      "ms",     "< 1500ms",  "PASS" if http_dur.get('p(90)', 0) < 1500 else "FAIL"),
            ("p(95) Response Time",     round(http_dur.get('p(95)', 0), 2),                      "ms",     "< 2000ms",  "PASS" if http_dur.get('p(95)', 0) < 2000 else "FAIL"),
            ("Error Rate",              round(error_rate_val * 100, 2),                          "%",      "< 10%",     "PASS" if error_rate_val < 0.1 else "FAIL"),
            ("",                        "",                                                       "",       "",          ""),
            ("Check Pass Rate",         round(checks.get('rate', 0) * 100, 2) if checks else 0, "%",      "> 95%",     "PASS" if checks.get('rate', 0) > 0.95 else "FAIL"),
            ("Data Received",           round(data_recv.get('count', 0) / (1024*1024), 2),       "MB",     "—",         "INFO"),
            ("Data Sent",               round(data_sent_m.get('count', 0) / (1024*1024), 2),     "MB",     "—",         "INFO"),
            ("Avg Connection Time",     round(http_conn.get('avg', 0), 2),                       "ms",     "< 100ms",   "PASS" if http_conn.get('avg', 0) < 100 else "FAIL"),
            ("Avg TLS Handshake",       round(http_tls.get('avg', 0), 2),                        "ms",     "< 200ms",   "PASS" if http_tls.get('avg', 0) < 200 else "FAIL"),
            ("Avg Waiting (TTFB)",      round(http_wait.get('avg', 0), 2),                       "ms",     "< 400ms",   "PASS" if http_wait.get('avg', 0) < 400 else "FAIL"),
            ("Total Iterations",        int(iterations.get('count', 0)),                         "count",  "—",         "INFO"),
            ("Iterations per Second",   round(iterations.get('rate', 0), 2),                     "iter/s", "> 5",       "PASS" if iterations.get('rate', 0) > 5 else "FAIL"),
        ]

        # Per-endpoint from custom metrics
        endpoint_prefixes = {
            'vehicles_req_duration':       'GET /vehicles',
            'users_req_duration':          'GET /users',
            'bookings_req_duration':       'GET /bookings',
            'notifications_req_duration':  'GET /notifications',
            'payments_req_duration':       'GET /payments',
            'reviews_req_duration':        'GET /reviews',
            'health_req_duration':         'Health Check',
        }

        endpoints = []
        for metric_key, display_name in endpoint_prefixes.items():
            m = metrics.get(metric_key, {}).get('values', {})
            if m:
                avg = round(m.get('avg', 0), 2)
                status = "OK" if avg < 1000 else "SLOW"
                endpoints.append((
                    display_name, "GET",
                    int(m.get('count', 0)) if 'count' in m else 'N/A',
                    avg,
                    round(m.get('min', 0), 2),
                    round(m.get('max', 0), 2),
                    status
                ))

        return kpis, endpoints, "Live K6 Run"

    except (json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
        print(f"[WARN] Could not parse K6 JSON ({e}), using simulated baseline data.")
        return None


# ═══════════════════════════════════════════════════════════════════════════
# Excel Report Generator
# ═══════════════════════════════════════════════════════════════════════════

def generate_excel(kpis, endpoints, config, data_source, output_path):
    """Generate a richly styled Excel report matching the original format."""
    wb = openpyxl.Workbook()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # ── Styles ────────────────────────────────────────────────────────────
    dark_bg       = PatternFill("solid", fgColor="1B1F2A")
    title_fill    = PatternFill("solid", fgColor="6C63FF")
    header_fill   = PatternFill("solid", fgColor="3949AB")
    metric_fill   = PatternFill("solid", fgColor="232836")
    pass_fill     = PatternFill("solid", fgColor="1B5E20")
    fail_fill     = PatternFill("solid", fgColor="B71C1C")
    info_fill     = PatternFill("solid", fgColor="263238")
    ep_title_fill = PatternFill("solid", fgColor="00695C")
    cfg_title_fill = PatternFill("solid", fgColor="E65100")

    white_font    = Font(color="FFFFFF", size=11)
    bold_white    = Font(color="FFFFFF", size=11, bold=True)
    header_font   = Font(color="FFFFFF", size=12, bold=True)
    title_font    = Font(color="FFFFFF", size=16, bold=True)
    sub_font      = Font(color="B0BEC5", size=10, italic=True)
    value_font    = Font(color="E0E0E0", size=11)
    kpi_name_font = Font(color="90CAF9", size=11, bold=True)
    pass_font     = Font(color="69F0AE", size=11, bold=True)
    fail_font     = Font(color="FF5252", size=11, bold=True)
    info_font     = Font(color="B0BEC5", size=11)
    green_val     = Font(color="00E676", size=12, bold=True)
    ok_font       = Font(color="69F0AE", size=11, bold=True)

    thin_border = Border(
        left=Side(style='thin', color='37474F'),
        right=Side(style='thin', color='37474F'),
        top=Side(style='thin', color='37474F'),
        bottom=Side(style='thin', color='37474F'),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left", vertical="center")

    def fill_dark(ws, rows, cols):
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c).fill = dark_bg

    def styled_cell(ws, r, c, val, font=white_font, fill=metric_fill, align=left_a, border=thin_border):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = border
        return cell

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: Executive Summary
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = "6C63FF"
    fill_dark(ws1, 35, 7)

    # Title row
    ws1.merge_cells('A1:F1')
    c = ws1['A1']
    c.value = "AgriRent — K6 Baseline Load Test Report"
    c.font = title_font
    c.fill = title_fill
    c.alignment = center

    # Metadata rows
    meta = [
        ("Report Generated",   now),
        ("Data Source",         data_source),
        ("Test Duration",       "1 minute (10s ramp-up / 40s hold / 10s ramp-down)"),
        ("Virtual Users (Peak)", 100),
        ("Target API",          "Supabase REST — atafeyidjdzedbktzivx.supabase.co"),
    ]
    for i, (label, val) in enumerate(meta, start=3):
        styled_cell(ws1, i, 1, label, font=bold_white, fill=metric_fill, align=left_a)
        styled_cell(ws1, i, 2, val, font=value_font, fill=metric_fill, align=left_a)

    # Section header: Key Performance Indicators
    kpi_start = 10
    ws1.merge_cells(f'A{kpi_start}:F{kpi_start}')
    sec = ws1[f'A{kpi_start}']
    sec.value = "Key Performance Indicators"
    sec.font = Font(color="FFFFFF", size=13, bold=True)
    sec.fill = PatternFill("solid", fgColor="37474F")
    sec.alignment = center

    # KPI table header
    kpi_headers = ["KPI", "Value", "Unit", "Threshold", "Result"]
    for ci, h in enumerate(kpi_headers, 1):
        styled_cell(ws1, kpi_start + 1, ci, h, font=header_font, fill=header_fill, align=center)

    # KPI data rows
    row = kpi_start + 2
    for kpi_name, val, unit, threshold, result in kpis:
        if kpi_name == "":
            row += 1
            continue
        styled_cell(ws1, row, 1, kpi_name, font=kpi_name_font, fill=metric_fill, align=left_a)
        styled_cell(ws1, row, 2, val, font=green_val, fill=metric_fill, align=center)
        styled_cell(ws1, row, 3, unit, font=value_font, fill=metric_fill, align=center)
        styled_cell(ws1, row, 4, threshold, font=value_font, fill=metric_fill, align=center)

        # Result cell color
        if result == "PASS":
            styled_cell(ws1, row, 5, result, font=pass_font, fill=pass_fill, align=center)
        elif result == "FAIL":
            styled_cell(ws1, row, 5, result, font=fail_font, fill=fail_fill, align=center)
        else:
            styled_cell(ws1, row, 5, result, font=info_font, fill=info_fill, align=center)
        row += 1

    # Column widths
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 5

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: Endpoint Breakdown
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Endpoint Breakdown")
    ws2.sheet_properties.tabColor = "00BCD4"
    fill_dark(ws2, len(endpoints) + 55, 8)

    # Title
    ws2.merge_cells('A1:G1')
    t2 = ws2['A1']
    t2.value = "Per-Endpoint Load Test Results"
    t2.font = title_font
    t2.fill = ep_title_fill
    t2.alignment = center

    # Headers
    ep_headers = ["Endpoint", "Method", "Requests", "Avg(ms)", "Min(ms)", "Max(ms)", "Status"]
    for ci, h in enumerate(ep_headers, 1):
        styled_cell(ws2, 2, ci, h, font=header_font, fill=header_fill, align=center)

    # Endpoint data rows
    r = 3
    for ep_name, method, reqs, avg, mn, mx, status in endpoints:
        styled_cell(ws2, r, 1, ep_name, font=bold_white, fill=metric_fill, align=left_a)
        styled_cell(ws2, r, 2, method, font=value_font, fill=metric_fill, align=center)
        styled_cell(ws2, r, 3, reqs, font=value_font, fill=metric_fill, align=center)
        styled_cell(ws2, r, 4, avg, font=green_val, fill=metric_fill, align=center)
        styled_cell(ws2, r, 5, mn, font=value_font, fill=metric_fill, align=center)
        styled_cell(ws2, r, 6, mx, font=value_font, fill=metric_fill, align=center)
        styled_cell(ws2, r, 7, status, font=ok_font, fill=metric_fill, align=center)
        r += 1

    # Column widths
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 12

    # ── Bar Chart: Average Response Time per Endpoint ────────────────────
    if endpoints:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Average Response Time per Endpoint (ms)"
        chart.y_axis.title = "Response Time (ms)"
        chart.x_axis.title = "Endpoint"
        chart.style = 10
        chart.width = 24
        chart.height = 14

        # Data reference = Avg(ms) column (col 4), from header row (2) to last data row
        data_ref = Reference(ws2, min_col=4, min_row=2, max_row=2 + len(endpoints))
        cats_ref = Reference(ws2, min_col=1, min_row=3, max_row=2 + len(endpoints))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4

        # Color each bar differently
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
        colors = ["4CAF50", "2196F3", "FF9800", "9C27B0", "F44336",
                  "00BCD4", "795548", "607D8B", "E91E63", "3F51B5"]
        series = chart.series[0]
        for idx in range(len(endpoints)):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = colors[idx % len(colors)]
            series.data_points.append(pt)

        ws2.add_chart(chart, f"A{r + 2}")

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: Test Configuration
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Test Configuration")
    ws3.sheet_properties.tabColor = "FF9800"
    fill_dark(ws3, 25, 4)

    # Title
    ws3.merge_cells('A1:B1')
    t3 = ws3['A1']
    t3.value = "Load Test Configuration"
    t3.font = title_font
    t3.fill = cfg_title_fill
    t3.alignment = center

    # Config rows
    row = 3
    for label, val in config:
        styled_cell(ws3, row, 1, label, font=bold_white, fill=metric_fill, align=left_a)
        styled_cell(ws3, row, 2, val, font=value_font, fill=metric_fill, align=left_a)
        row += 1

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 55

    # ── Save ──────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    print(f"[OK] K6 Load Test Excel report generated at: {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    json_path = sys.argv[1] if len(sys.argv) > 1 else "test_reports/k6_results.json"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "test_reports/K6_Load_Test_Report.xlsx"

    # Try to use real K6 data; fall back to simulated baseline
    parsed = try_parse_k6_json(json_path)

    if parsed:
        kpis, endpoints, data_source = parsed
        print(f"[INFO] Using live K6 data from: {json_path}")
    else:
        kpis = get_baseline_kpis()
        endpoints = get_baseline_endpoints()
        data_source = "Simulated Baseline"
        print("[INFO] Using simulated baseline data (no valid K6 JSON found).")

    config = get_baseline_config()
    generate_excel(kpis, endpoints, config, data_source, output_path)
