"""
.github/scripts/csv_to_excel.py
Converts AgriRent_Full_Functionality_Test_Report.csv -> test/AgriRent_Test_Report.xlsx
Run by GitHub Actions workflow: .github/workflows/generate_test_report.yml
"""

import csv
import os
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
COL_HEADER_BG  = "1B5E20"
COL_HEADER_FG  = "FFFFFF"
COL_PASS       = "E8F5E9"
COL_FAIL       = "FFEBEE"
COL_WARN       = "FFF8E1"
COL_COND       = "E3F2FD"
COL_PASS_BADGE = "2E7D32"
COL_FAIL_BADGE = "C62828"
COL_WARN_BADGE = "F57F17"
COL_COND_BADGE = "1565C0"
COL_ALT_ROW    = "F5F5F5"


def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def row_fill(status: str, idx: int):
    s = status.upper()
    if "FAIL" in s:        return PatternFill("solid", fgColor=COL_FAIL)
    if "WARN" in s:        return PatternFill("solid", fgColor=COL_WARN)
    if "CONDITIONAL" in s: return PatternFill("solid", fgColor=COL_COND)
    if "PASS" in s:        return PatternFill("solid", fgColor=COL_PASS)
    return PatternFill("solid", fgColor=COL_ALT_ROW if idx % 2 == 0 else "FFFFFF")


def status_font_color(status: str):
    s = status.upper()
    if "FAIL" in s:        return COL_FAIL_BADGE
    if "WARN" in s:        return COL_WARN_BADGE
    if "CONDITIONAL" in s: return COL_COND_BADGE
    if "PASS" in s:        return COL_PASS_BADGE
    return "000000"


def severity_fill(sev: str):
    s = sev.upper()
    colors = {"CRITICAL": "B71C1C", "HIGH": "E65100",
              "MEDIUM": "F57F17", "LOW": "F9A825"}
    c = colors.get(s)
    return PatternFill("solid", fgColor=c) if c else None


def severity_font(sev: str):
    s = sev.upper()
    bold  = s in ("CRITICAL", "HIGH")
    color = "FFFFFF" if s in ("CRITICAL", "HIGH") else "000000"
    return Font(bold=bold, color=color, size=10)


def write_summary_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.fill      = PatternFill("solid", fgColor=COL_HEADER_BG)
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border()
    return c


def write_summary_cell(ws, row, col, value, bold=False, color="000000", bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=color, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


def main():
    # Paths are relative to repo root (where the action runs)
    CSV_PATH  = "AgriRent_Full_Functionality_Test_Report.csv"
    XLSX_PATH = "test/AgriRent_Test_Report.xlsx"

    os.makedirs("test", exist_ok=True)

    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    headers = rows[0]
    data    = rows[1:]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Report"

    # Column widths
    col_widths = {1: 10, 2: 14, 3: 22, 4: 42, 5: 13, 6: 18, 7: 12, 8: 58, 9: 48}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    # Header row
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = PatternFill("solid", fgColor=COL_HEADER_BG)
        cell.font      = Font(bold=True, color=COL_HEADER_FG, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[1].height = 30

    STATUS_COL   = 6
    SEVERITY_COL = 7

    # Data rows
    for ri, row in enumerate(data, start=2):
        status   = row[STATUS_COL - 1]   if len(row) >= STATUS_COL   else ""
        severity = row[SEVERITY_COL - 1] if len(row) >= SEVERITY_COL else ""
        fill     = row_fill(status, ri)

        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = Alignment(
                vertical="top", wrap_text=True,
                horizontal="center" if ci in (1, 5, 6, 7) else "left"
            )
            cell.font = Font(size=10)

            if ci == STATUS_COL:
                cell.font = Font(bold=True, color=status_font_color(status), size=10)

            if ci == SEVERITY_COL:
                sf = severity_fill(severity)
                if sf:
                    cell.fill = sf
                cell.font      = severity_font(severity)
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        ws.row_dimensions[ri].height = 55

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    status_counts   = Counter(r[STATUS_COL - 1]   for r in data if len(r) >= STATUS_COL)
    severity_counts = Counter(r[SEVERITY_COL - 1] for r in data if len(r) >= SEVERITY_COL)
    module_counts   = Counter(r[1]                for r in data if len(r) > 1)

    ws2.merge_cells("A1:D1")
    t = ws2["A1"]
    t.value     = "AgriRent - Test Report Summary"
    t.font      = Font(bold=True, size=14, color=COL_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    # By Status
    write_summary_header(ws2, 3, 1, "Status")
    write_summary_header(ws2, 3, 2, "Count")
    write_summary_header(ws2, 3, 3, "% of Total")
    total = sum(status_counts.values())
    ri = 4
    for s in ["PASS", "FAIL", "WARN", "CONDITIONAL PASS"]:
        cnt = status_counts.get(s, 0)
        pct = f"{cnt/total*100:.0f}%" if total else "0%"
        bg  = {"PASS": COL_PASS, "FAIL": COL_FAIL,
               "WARN": COL_WARN, "CONDITIONAL PASS": COL_COND}.get(s)
        write_summary_cell(ws2, ri, 1, s, bold=True, color=status_font_color(s), bg=bg)
        write_summary_cell(ws2, ri, 2, cnt, bold=True)
        write_summary_cell(ws2, ri, 3, pct)
        ri += 1
    write_summary_cell(ws2, ri, 1, "TOTAL", bold=True, bg="E0E0E0")
    write_summary_cell(ws2, ri, 2, total,  bold=True, bg="E0E0E0")
    write_summary_cell(ws2, ri, 3, "100%",             bg="E0E0E0")

    # By Severity
    ri += 2
    write_summary_header(ws2, ri, 1, "Severity")
    write_summary_header(ws2, ri, 2, "Count")
    ri += 1
    for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]:
        cnt  = severity_counts.get(s, 0)
        sf   = severity_fill(s)
        bg   = sf.fgColor.value if sf else "FFFFFF"
        fc   = severity_font(s).color.value
        write_summary_cell(ws2, ri, 1, s, bold=True, color=fc, bg=bg)
        write_summary_cell(ws2, ri, 2, cnt, bold=True)
        ri += 1

    # By Module
    ri += 2
    write_summary_header(ws2, ri, 1, "Module")
    write_summary_header(ws2, ri, 2, "Test Cases")
    ri += 1
    for mod, cnt in sorted(module_counts.items()):
        write_summary_cell(ws2, ri, 1, mod)
        write_summary_cell(ws2, ri, 2, cnt, bold=True)
        ri += 1

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    wb.save(XLSX_PATH)
    print(f"Saved: {XLSX_PATH}  ({os.path.getsize(XLSX_PATH):,} bytes)")
    print(f"Rows : {len(data)}")
    print(f"Sheets: Test Report, Summary")


if __name__ == "__main__":
    main()
